"""
    Saves the finished bill of material file.
"""

from pathlib import Path
from typing import List

from const import BOM as BOM_FOLDER
from helper.resource import ResourceCommons
from models.bom import BOM
from models.bom import BOMAssemblyItem
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from protocols.task_protocol import TaskProtocol
from pytia.log import log
from resources import resource
from utils.excel import create_header
from utils.excel import style_worksheet
from utils.excel import write_data


class SaveBomTask(TaskProtocol):
    """
    This class is used to format and save the finished bill of material file.

    Args:
        TaskProtocol (_type_): The task protocol.
    """

    __slots__ = ("_bom", "_path")

    def __init__(self, bom: BOM, export_root_path: Path, filename: str) -> None:
        self.bom = bom
        self.export_root_path = export_root_path
        self.filename = filename

    def run(self) -> None:
        """Runs the task."""
        log.info("Saving finished bill of material.")
        self._save_bom(
            bom=self.bom,
            folder=Path(self.export_root_path, BOM_FOLDER),
            filename=self.filename,
        )

    @classmethod
    def _save_bom(cls, bom: BOM, folder: Path, filename: str) -> None:
        """
        Saves the bill of material from the BOM object as xlsx file. 
        This saves only the content of the BOM object, regardless of wether the Report 
        is OK or FAILED.

        Args:
            bom (BOM): The BOM object to save.
            folder (Path): The path into which to save the bill of material.
            filename (str): The name of the file to save. '.xlsx' will be added if not \
                in name. Separate files will be created if set in bom.json.
        """

        wb_summary = Workbook()
        ws_summary: Worksheet = wb_summary.active  # type: ignore
        ws_summary.title = "Summary"

        wb_made: Workbook | None = None
        wb_bought: Workbook | None = None

        if resource.bom.files.separate:
            wb_made = Workbook()
            ws_made: Worksheet = wb_made.active  # type: ignore
            ws_made.title = "Made"

            wb_bought = Workbook()
            ws_bought: Worksheet = wb_bought.active  # type: ignore
            ws_bought.title = "Bought"
        else:
            ws_made = wb_summary.create_sheet(title="Made")
            ws_bought = wb_summary.create_sheet(title="Bought")

        # Summary
        cls._write_worksheet(
            worksheet=ws_summary,
            header=resource.bom.header_items.summary,
            data=bom.summary.items,
            strict=True,
        )

        # Made
        if resource.bom.header_items.made:
            cls._write_worksheet(
                worksheet=ws_made,
                header=resource.bom.header_items.made,
                data=[
                    item
                    for item in bom.summary.items
                    if item.source == resource.applied_keywords.made
                ],
                strict=False,
            )

        # Bought
        if resource.bom.header_items.bought:
            cls._write_worksheet(
                worksheet=ws_bought,
                header=resource.bom.header_items.bought,
                data=[
                    item
                    for item in bom.summary.items
                    if item.source == resource.applied_keywords.bought
                ],
                strict=False,
            )

        for assembly in bom.assemblies:
            ws_assembly: Worksheet = wb_summary.create_sheet(title=assembly.partnumber)
            cls._write_worksheet(
                worksheet=ws_assembly,
                header=resource.bom.header_items.summary,
                data=assembly.items,
                strict=True,
            )

        wb_summary.active = wb_summary["Summary"]

        # File operations
        if ".xlsx" in filename:
            filename = filename.split(".xlsx")[0]

        if resource.bom.files.separate:
            path_sum = Path(folder, f"{filename} ({resource.bom.files.summary}).xlsx")
            path_made = Path(folder, f"{filename} ({resource.bom.files.made}).xlsx")
            path_bought = Path(folder, f"{filename} ({resource.bom.files.bought}).xlsx")

            if wb_made and resource.bom.header_items.made:
                wb_made.save(str(path_made))

            if wb_bought and resource.bom.header_items.bought:
                wb_bought.save(str(path_bought))
        else:
            path_sum = Path(folder, filename + ".xlsx")

        wb_summary.save(str(path_sum))
        log.info(f"Saved processed BOM to {str(folder)!r}.")

    @staticmethod
    def _write_worksheet(
        worksheet: Worksheet,
        header: list,
        data: List[BOMAssemblyItem],
        strict: bool,
    ) -> None:
        create_header(worksheet=worksheet, header_items=header)
        write_data(
            worksheet=worksheet,
            header_items=header,
            data=data,
            strict=strict,
        )
        style_worksheet(worksheet=worksheet)
