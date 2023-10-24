"""
    Processes the xlsx file and generates the BOM object from it.
"""

import re
from pathlib import Path
from typing import Any
from typing import Dict

from const import KEEP
from const import X000D
from helper.commons import ResourceCommons
from models.bom import BOM
from models.bom import BOMAssembly
from models.bom import BOMAssemblyItem
from models.paths import Paths
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from protocols.task_protocol import TaskProtocol
from pytia.log import log
from resources import resource
from utils.excel import row_is_empty


class ProcessBomTask(TaskProtocol):
    """
    This class processes the bill of material from the converted xlsx file.
    It generates the BOM object from the xlsx file.

    Args:
        TaskProtocol (_type_): The protocol for the task runner.
    """

    __slots__ = (
        "_xlsx",
        "_project",
        "_paths",
        "_bom",
        "_ignore_prefix",
        "_ignore_prefix_txt",
        "_ignore_bought",
    )

    def __init__(
        self,
        xlsx: Path,
        project_number: str,
        paths: Paths,
        ignore_prefix_txt: str | None,
        ignore_source_unknown: bool,
    ) -> None:
        self._xlsx = xlsx
        self._project = project_number
        self._paths = paths
        self._ignore_prefix_txt = ignore_prefix_txt
        self._ignore_source_unknown = ignore_source_unknown

    @property
    def bom(self) -> BOM:
        return self._bom

    def run(self) -> None:
        """Runs the task."""
        log.info("Processing bill of material.")

        wb = self._get_workbook_from_xlsx(xlsx_path=self._xlsx)
        self._bom = self._retrieve_bom_from_catia_export(
            worksheet=wb.worksheets[0],
            overwrite_project=self._project,
        )
        self._sort_bom(bom=self._bom)

    @staticmethod
    def _get_workbook_from_xlsx(xlsx_path: Path) -> Workbook:
        """
        Returns a Workbook object from the xlsx file. Closes the file afterwards. Only reads the data.

        Args:
            xlsx_path (Path): The path to the xlsx file.

        Returns:
            Workbook: The workbook.
        """
        workbook = load_workbook(filename=str(xlsx_path), data_only=True)
        workbook.close()
        return workbook

    def _retrieve_bom_from_catia_export(
        self,
        worksheet: Worksheet | ReadOnlyWorksheet,
        overwrite_project: str = KEEP,
    ) -> BOM:
        """
        Extracts all bill of materials and the summary from the CATIA export EXCEL file.
        Some keyword names depend on the language of the CATIA UI. It is therefor necessary to
        check the language first, the the keywords can be matched.

        Args:
            worksheet (Worksheet | ReadOnlyWorksheet): The worksheet from which to extract the data.
            paths (Paths): The paths from all CATIA partnumbers in the BOM.
            overwrite_project (str): The project number that will be written into the BOM object. \
                If set to `KEEP` all existing project number will be left, otherwise the provided 
                string will be written into the BOM object.

        Raises:
            KeyError: Raised when the exported EXCEL file has no partnumber column.
            KeyError: Raised when there is a partnumber in the BOM that is not in the paths list.

        Returns:
            BOM: The bill of material object.
        """
        bom = BOM()
        assembly: BOMAssembly | None = None
        header_positions: dict = {}
        data_row = 0
        is_summary = False

        # Do not use worksheet.max_row during the iteration over the worksheet.
        # Worksheet.max_row will change during runtime, when accessing a row greater
        # than the current max_row.
        max_row = worksheet.max_row

        log.info("Retrieving bill of material from exported Excel file.")

        # The bill of material will be processed from the summary-header-items.
        # At this point the made and bought header items aren't relevant, those come
        # into play only at the final step of saving the bill of material later.
        # Also: The header items are defined as HEADER_NAME:PROPERTY_NAME (see bom.json).
        # At this it's necessary to extract the PROPERTY_NAME first, to create the
        # correct output, because CATIA needs the property names to create the bill of
        # material.
        header_items = ResourceCommons.get_property_names_from_config(
            resource.bom.header_items.summary
        )

        for ri in range(1, max_row + 2):
            log.debug(f"Working on row {ri} of {worksheet.max_row}.")

            # Since we iterate over the whole catia export excel file, and this export
            # contains all bills of material (all boms from all sub-assemblies and the
            # bom-summary from the whole thing), we need to identify where we are in the
            # current row. This is done with the _bom_or_summary variable and the _name
            # variable.
            # The _bom_or_summary variable defines wether the current row is
            # a bom from a sub-assembly or from the summary.
            # The _name variable defines the name (partnumber) of the parent product.
            _bom_or_summary = str(worksheet.cell(ri, 1).value).split(": ")[0]
            _name = str(worksheet.cell(ri, 1).value).split(": ")[-1]

            # An empty row means that the either the bom hasn't begun or it has ended.
            # So we need to check if the assembly object isn't none (and if we are not
            # at the summary section) to determine if we can add the current assembly
            # data to the list of assemblies.
            # The assembly object needs to be set to none after adding its content to
            # the list, so this condition isn't reached by accident.
            # In short: This if condition is for adding a sub-assembly to the list of
            # assemblies.
            if row_is_empty(worksheet, ri) and assembly is not None and not is_summary:
                bom.assemblies.append(assembly)
                log.debug(
                    f"Added assembly of {assembly.partnumber!r} to the list of assemblies."
                )
                assembly = None

            # Same as above, but we check here if the assembly object is the summary.
            # The summary data-block comes always last in the catia export excel file,
            # at this we can just check if we're at the end of the excel file (in case
            # you wondered why we iterate over worksheet.max_row + 2, this is why).
            elif ri > max_row and assembly is not None and is_summary:
                bom.summary = assembly
                log.debug(f"Added assembly of {assembly.partnumber!r} to the summary.")

            # To check if a new bom needs to be processed we need to check if the
            # keyword "Bill of Material" is in the current row.
            # If that is the case we create a new assembly object and check the header
            # of said new assembly. Those headers should always be those from the
            # bom.json file (but checking is better than hoping).
            elif resource.applied_keywords.bom in _bom_or_summary:
                assembly = BOMAssembly(partnumber=_name, path=self._paths.items[_name])
                header_positions = self._get_header_positions(
                    worksheet=worksheet, row=ri + 1, header_items=header_items
                )
                # The data row for sub-assembly-boms is always two rows after the
                # "Bill of Material" keyword.
                data_row = ri + 2
                log.info(f"Processing BOM of element {_name!r}.")

            # To check if a new summary needs to be processed we need to check if the
            # keyword "Recapitulation" is in the current row.
            # If that is the case we create a new assembly object.
            # Note: The summary is always the last data-block of the excel file, so we
            # don't need to bother setting it back to false somewhere else.
            elif resource.applied_keywords.summary in _bom_or_summary:
                is_summary = True
                assembly = BOMAssembly(partnumber=_name, path=self._paths.items[_name])
                header_positions = self._get_header_positions(
                    worksheet=worksheet, row=ri + 4, header_items=header_items
                )
                # The data row for summary-boms is always five rows after the
                # "Bill of Material" keyword.
                data_row = ri + 5
                log.info(f"Processing BOM summary.")

            # Only when the assembly object isn't none (valid keyword "Bill of
            # Material" or "Recapitulation" and valid header positions) and the current
            # row number matches the beginning of the data row from the excel file we
            # can start treating the current row as data row.
            # All if-conditions before were only for validating the beginning or the end
            # of a bom-data-range.
            if (
                assembly is not None
                and ri >= data_row
                and not row_is_empty(worksheet, ri)
            ):
                # The row data dict will contain the data from the current row (wow).
                # It's keys are the header items names (from the bom.json) and it's
                # values are the actual data.
                row_data: dict = {}

                # To make sure we gather the right data from the right header we
                # use the header position dict to access the column of the excel file.
                # We don't iterate over all excel columns.
                for position in header_positions:
                    cell_value = worksheet.cell(ri, header_positions[position]).value

                    # This is a result of legacy catia macros. This isn't needed if all
                    # parts and products are setup with the pytia-property-manager.
                    if isinstance(cell_value, str) and X000D in cell_value:
                        cell_value = cell_value.replace("_x000D_\n", "\n")

                    # catia exports empty cells as chr(13). This results in a space
                    # string " " instead of a truly empty cell. This is fixed by
                    # checking for only-whitespace characters and replacing them with
                    # None.
                    if isinstance(cell_value, str) and re.match(r"^\s+$", cell_value):
                        cell_value = None

                    cell_value = self._overwrite_project_number(
                        header_position=position,
                        cell_value=cell_value,
                        overwrite_number=overwrite_project,
                    )

                    cell_value = self._translate_username(
                        header_position=position, cell_value=cell_value
                    )

                    cell_value = self._apply_fixed_text(
                        header_position=position, cell_value=cell_value
                    )

                    # cell_value = self._apply_placeholder_header(
                    #     header_position=position, cell_value=cell_value
                    # )

                    self._add_to_row_data(
                        row_data, header_position=position, cell_value=cell_value
                    )

                # The partnumber must be available in the header. We use the partnumber
                # to identify a part or product in the assembly.
                if not resource.applied_keywords.partnumber in row_data:
                    raise KeyError(
                        f"Cannot find keyword {resource.applied_keywords.partnumber!r} in exported "
                        "Excel file. Are your language settings correct? Or is the $partnumber "
                        "keyword not set in the bom.json's header_items?"
                    )

                # The partnumber must not be empty (this should be impossible).
                if row_data[resource.applied_keywords.partnumber] is None:
                    raise ValueError("The value for the partnumber is empty.")

                # Warn the user when a part or product is missing. This happens mostly
                # when there are items in the catia tree, that aren't stored in a file.
                # (Cameras, Simulations, etc.)
                if (
                    row_data[resource.applied_keywords.partnumber]
                    not in self._paths.items
                ):
                    item_path = None
                    log.warning(
                        "No path found for item "
                        f"{row_data[resource.applied_keywords.partnumber]!r}."
                    )
                else:
                    item_path = self._paths.items[
                        row_data[resource.applied_keywords.partnumber]
                    ]

                # Finally we check if the assembly item isn't tagged to be ignored.
                # If not, it's added to the dataclass.
                if (
                    self._ignore_source_unknown
                    and row_data[resource.applied_keywords.source]
                    == resource.applied_keywords.unknown
                ) or (
                    self._ignore_prefix_txt is not None
                    and any(
                        [
                            str(
                                row_data[resource.applied_keywords.partnumber]
                            ).startswith(s)
                            for s in self._ignore_prefix_txt.split(";")
                        ]
                    )
                ):
                    log.info(
                        " - Ignoring item "
                        f"{row_data[resource.applied_keywords.partnumber]!r}."
                    )
                else:
                    assembly_item = BOMAssemblyItem(
                        partnumber=row_data[resource.applied_keywords.partnumber],
                        source=row_data[resource.applied_keywords.source],
                        properties=row_data,
                        path=item_path,
                    )
                    assembly.items.append(assembly_item)
                    log.info(
                        f" - Added item {row_data[resource.applied_keywords.partnumber]!r} to element "
                        f"{assembly.partnumber!r}{' (summary).' if is_summary else '.'}"
                    )
        return bom

    @staticmethod
    def _get_header_positions(
        worksheet: Worksheet | ReadOnlyWorksheet, row: int, header_items: tuple
    ) -> Dict[str, int]:
        """
        Returns a dictionary of the header positions from the from CATIA exported raw bill of material.
        Maps the header positions from the bill of material to the header names.

        Example:
        The header of the raw export looks like this: `Project | Part Number | Revision`
        Then the returned dictionary looks like this: `{"Project": 1, "Part Number": 2, "Revision": 3}`

        The index starts with `1`, so the dictionary is usable on excel columns.

        This method is purely for verifying the raw CATIA export. If everything is correct, then
        CATIA will export all later required header items on the correct position, given in the
        bom.json config file.

        Args:
            worksheet (Worksheet | ReadOnlyWorksheet): The worksheet to get the header positions from.
            row (int): The row in which to look for the header positions.
            header_items (tuple): The header names to match.

        Raises:
            KeyError: Raised if the header names do not exist in the worksheet.

        Returns:
            Dict[str, int]: The dictionary of header positions.
        """
        header: dict = {}

        for i in range(1, worksheet.max_column + 1):
            header[worksheet.cell(row, i).value] = i

        # Check if CATIA exported all headers from the provided header items dict from the bom.json
        for item in header_items:
            if item not in header.keys():
                raise KeyError(
                    f"Header item {item!r} not found in exported bill of material."
                )

        # Check if all required header items from the bom.json are in the exported headers
        # TODO: Put this into another function, this function has too many responsibilities.
        for required_header in resource.bom.required_header_items.values:
            if required_header not in header.keys():
                raise KeyError(
                    f"The required header item {required_header!r} is not present in the exported "
                    "bill of material. Please add this header to the bom.json."
                )

        log.debug(f"Retrieved header items from excel worksheet from row {row}.")
        return header

    @staticmethod
    def _add_to_row_data(
        row_data: dict, header_position: str, cell_value: str | Any | None
    ) -> None:
        """
        Adds the cell value to the correct header position of the row_data dict.

        Args:
            row_data (dict): The row_data dict.
            header_position (str): The name of the header item (the key in the row_data)
            cell_value (str | Any | None): The value to add (the value in the row_data)
        """
        # if header_position.startswith("%") and "=" in header_position:
        #     row_data[header_position.split("%")[-1].split("=")[0]] = cell_value
        # else:
        #     row_data[header_position] = cell_value
        row_data[header_position] = cell_value

    @staticmethod
    def _overwrite_project_number(
        header_position: str, cell_value: str | Any | None, overwrite_number: str
    ) -> str | Any | None:
        """
        Returns the overwritten project number if the header position is the project and if the user
        selected the project number to be overwritten.
        Otherwise returns the given cell_value without changes.

        Args:
            header_position (str): The header position (the name of the header).
            cell_value (str | Any | None): The cell value to be overwritten.
            overwrite_number (str): The project number that will be written.

        Returns:
            str | Any | None: The cell value or the overwritten project number.
        """
        if (
            header_position == resource.bom.required_header_items.project
            and overwrite_number != KEEP
        ):
            return overwrite_number
        return cell_value

    @staticmethod
    def _translate_username(
        header_position: str, cell_value: str | Any | None
    ) -> str | Any | None:
        """
        Returns the translated username if the header position is creator or modifier.
        Otherwise returns the given cell_value without changes.

        Args:
            header_position (str): The header position (the name of the header).
            cell_value (str | Any | None): The cell value to be translated.

        Returns:
            str | Any | None: The translated username or the cell value.
        """
        if (
            resource.settings.export.apply_username_in_bom
            and (
                header_position == resource.props.creator
                or header_position == resource.props.modifier
            )
            and resource.user_exists(str(cell_value))
        ):
            return resource.get_user_by_logon(str(cell_value)).name
        return cell_value

    @staticmethod
    def _apply_fixed_text(
        header_position: str, cell_value: str | Any | None
    ) -> str | Any | None:
        """
        Returns the value of the fixed text accordingly to the header_items list.

        Args:
            header_position (str): The header position (the name of the header).
            cell_value (str | Any | None): The cell value.

        Returns:
            str | Any | None: The fixed text, if set in the header_items.
        """

        for item in resource.bom.header_items.summary:
            if header_position in item and "=" in item:
                return item.split("=")[-1]
        return cell_value

    # @staticmethod
    # def _apply_placeholder_header(
    #     header_position: str, cell_value: str | Any | None
    # ) -> str | Any | None:
    #     """
    #     Returns None as cell value, if the header is set to placeholder.

    #     Args:
    #         header_position (str): The header position (the name of the header).
    #         cell_value (str | Any | None): The cell value.

    #     Returns:
    #         str | Any | None: The fixed text, if set in the header_items.
    #     """
    #     if not any(delimiter in header_position for delimiter in [":", "="]):
    #         return None
    #     return cell_value

    @staticmethod
    def _sort_bom(bom: BOM) -> None:
        """
        Sorts the BOM model according to the `sort` item in the bom.json.
        Note: This depends on the language of the CATIA UI. Make sure that you have
        executed the `resources.apply_keywords_to_export()` method before calling this function.

        Args:
            bom (BOM): The BOM object to be sorted.
        """

        def _partnumber(assembly_item: BOMAssemblyItem):
            return assembly_item.partnumber

        def _made(assembly_item: BOMAssemblyItem):
            props = assembly_item.properties
            log.debug(f"Sorting 'made' items by {resource.bom.sort.made!r}")
            return (
                props[resource.bom.sort.made]
                if props[resource.applied_keywords.source]
                == resource.applied_keywords.made
                else None
            )

        def _bought(assembly_item: BOMAssemblyItem):
            props = assembly_item.properties
            log.debug(f"Sorting 'bought' items by {resource.bom.sort.bought!r}")
            return (
                props[resource.bom.sort.bought]
                if props[resource.applied_keywords.source]
                == resource.applied_keywords.bought
                else None
            )

        bom.summary.items.sort(key=lambda x: (_bought(x) is None, _bought(x)))
        bom.summary.items.sort(key=lambda x: (_made(x) is None, _made(x)))
        log.info("Sorted BOM items of 'Summary'.")

        for assembly in bom.assemblies:
            # assembly.items.sort(key=_source, reverse=True)
            assembly.items.sort(key=lambda x: (_bought(x) is None, _bought(x)))
            assembly.items.sort(key=lambda x: (_made(x) is None, _made(x)))
            log.debug(f"Sorted BOM items of {assembly.partnumber}.")
        log.info("Sorted BOM items of all assemblies.")
