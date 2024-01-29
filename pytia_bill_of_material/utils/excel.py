"""
    Excel utility: Functions for handling Excel files.
"""

from typing import List

from helper.resource import ResourceCommons
from models.bom import BOMAssemblyItem
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from pytia.exceptions import PytiaDispatchError
from pytia.exceptions import PytiaNotInstalledError
from pytia.log import log
from resources import resource
from win32com.client import CDispatch
from win32com.client import Dispatch
from win32com.server.exception import COMException


def get_excel() -> CDispatch:
    """
    Returns the EXCEL com object.

    Raises:
        PytiaNotInstalledError: Raised when MS Excel is not installed.
        PytiaDispatchError: Raises when the connection to the Excel com host fails.

    Returns:
        CDispatch: The Excel com object.
    """
    try:
        # app = EnsureDispatch("EXCEL.Application")
        app = Dispatch("EXCEL.Application")
        return app  # type: ignore
    except COMException as e:
        raise PytiaNotInstalledError(
            f"Excel is not installed on this system: {e}"
        ) from e
    except Exception as e:  # pylint: disable=broad-except
        raise PytiaDispatchError(f"Failed connecting to Excel: {e}") from e


def create_header(worksheet: Worksheet, header_items: list) -> None:
    """
    Creates a header row in the given worksheet.

    Args:
        worksheet (Worksheet): The worksheet into which to create the header row.
        header_items (tuple): The header items to create (the bom.json list).
    """
    items = ResourceCommons.get_header_names_from_config(header_items)

    if isinstance(resource.bom.header_row, int):
        for index, item in enumerate(items):
            worksheet.cell(resource.bom.header_row + 1, index + 1, str(item))
        log.info(f"Created header for worksheet {worksheet.title!r}")
    else:
        log.info(f"Skipped creating header for worksheet {worksheet.title!r}.")


def write_data(
    worksheet: Worksheet, header_items: list, data: List[BOMAssemblyItem], strict: bool
) -> None:
    """
    Writes the properties from the given data object to the given worksheet.
    Takes care that each datum will be written to the corresponding header.

    Args:
        worksheet (Worksheet): The worksheet to add the data to.
        header_items (tuple): The header items (the bom.json list).
        data (List[BOMAssemblyItem]): The actual data.
        strict (bool): If set to True, an error will be raised if the properties are \
            not in the header items.
    """
    items = ResourceCommons.get_property_names_from_config(header_items)

    for ri, rv in enumerate(data):
        for ci, cv in enumerate(items):
            if cv in rv.properties:
                cell_value = rv.properties[cv]
            elif strict:
                raise KeyError(f"Did not find property {cv!r} in data.")
            else:
                cell_value = None
                # log.warning(
                #     f"Did not find property {cv!r} in data (is this header missing in "
                #     "the header_items.summary list of the bom.json config file?"
                # )
            worksheet.cell(
                row=ri + resource.bom.data_row + 1, column=ci + 1
            ).value = cell_value
    log.info(f"Wrote all data to worksheet {worksheet.title!r}.")


def row_is_empty(worksheet: Worksheet | ReadOnlyWorksheet, row: int) -> bool:
    """Returns wether the given row is empty or not."""
    max_column = worksheet.max_column
    assert max_column
    return all([worksheet.cell(row, i).value is None for i in range(1, max_column + 1)])


def style_worksheet(worksheet: Worksheet) -> None:
    """Styles the worksheet as stated in the bom.json config file."""
    for column_cells in worksheet.columns:
        # Set format, font and size
        for index, cell in enumerate(column_cells):  # type: ignore
            if index > resource.bom.data_row - 1:
                color = (
                    resource.bom.data_color_1
                    if index % 2 == 0
                    else resource.bom.data_color_2
                )
                bg_color = (
                    resource.bom.data_bg_color_1
                    if index % 2 == 0
                    else resource.bom.data_bg_color_2
                )
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
            else:
                color = None
            cell.number_format = "@"
            cell.font = Font(
                name=resource.bom.font, size=resource.bom.size, color=color
            )
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Set font and height for the header row
        if isinstance(resource.bom.header_row, int):
            worksheet.row_dimensions[resource.bom.header_row + 1].height = 20  # type: ignore
            column_cells[resource.bom.header_row].font = Font(  # type: ignore
                name=resource.bom.font,
                size=resource.bom.size,
                bold=True,
                color=resource.bom.header_color,
            )
            column_cells[resource.bom.header_row].fill = PatternFill(  # type: ignore
                start_color=resource.bom.header_bg_color,
                end_color=resource.bom.header_bg_color,
                fill_type="solid",
            )
            column_cells[resource.bom.header_row].alignment = Alignment(  # type: ignore
                horizontal="center", vertical="center"
            )

        # Set cell width
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)  # type: ignore
        worksheet.column_dimensions[column_cells[0].column_letter].width = (  # type: ignore
            length if length > 2 else 2
        )

    log.info(f"Styled worksheet {worksheet.title!r}.")
